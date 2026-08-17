"""MarkItDown extraction lane — binary documents to plain text only.

Important: never mutate the caller's original file. Binary extract always runs on a
temp copy under the system temp directory.
"""

from __future__ import annotations

import shutil
import tempfile
from pathlib import Path

from duckclaw.document_toolbox.registry import EXTRACT_SUFFIXES, INGEST_NATIVE_SUFFIXES


def markitdown_available() -> bool:
    try:
        import markitdown  # noqa: F401

        return True
    except ImportError:
        return False


def _is_magika_model_error(exc: BaseException) -> bool:
    msg = str(exc).lower()
    return "model dir not found" in msg or ("magika" in msg and "model" in msg)


def _excel_to_text(path: Path) -> str:
    """Fallback when MarkItDown/Magika is unavailable (e.g. PyInstaller without models)."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError(
            "No se pudo leer Excel: falta openpyxl y Magika/MarkItDown no está disponible."
        ) from exc

    # read_only + data_only: never write back to the workbook on disk.
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        parts: list[str] = []
        for sheet in wb.worksheets:
            parts.append(f"## {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                cells = ["" if c is None else str(c).strip() for c in row]
                if any(cells):
                    parts.append(" | ".join(cells))
        return "\n".join(parts).strip()
    finally:
        wb.close()


def _fallback_extract_by_suffix(path: Path, *, reason: str) -> str:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        text = _excel_to_text(path)
        if text:
            return text
        raise ValueError(f"Excel vacío o sin celdas legibles: {path.name}")
    if suffix == ".csv":
        return path.read_text(encoding="utf-8", errors="replace").strip()
    raise ValueError(
        f"No se pudo extraer texto de {path.name} ({suffix or 'sin extensión'}). "
        f"Causa: {reason}"
    )


def _convert_path(path: Path) -> str:
    """Extract text from ``path``. Caller must pass a disposable temp copy for binaries."""
    suffix = path.suffix.lower()
    try:
        from markitdown import MarkItDown, StreamInfo
    except ImportError as exc:
        if suffix in {".xlsx", ".xlsm", ".csv"}:
            return _fallback_extract_by_suffix(path, reason="MarkItDown no instalado")
        raise ValueError(
            "MarkItDown no instalado. Ejecuta: uv sync (o duckops up)"
        ) from exc

    try:
        md = MarkItDown()
    except Exception as exc:
        if _is_magika_model_error(exc) and suffix in {".xlsx", ".xlsm", ".xls", ".csv"}:
            if suffix == ".xls":
                raise ValueError(
                    f"No se pudo leer .xls ({path.name}): Magika/modelos no disponibles "
                    "en este build. Guarda como .xlsx e inténtalo de nuevo."
                ) from exc
            return _fallback_extract_by_suffix(path, reason=str(exc))
        raise

    try:
        stream_info = StreamInfo(extension=suffix) if suffix else None
        result = md.convert(str(path), stream_info=stream_info)
        return (result.text_content or "").strip()
    except Exception as exc:
        if _is_magika_model_error(exc) and suffix in {".xlsx", ".xlsm", ".csv"}:
            return _fallback_extract_by_suffix(path, reason=str(exc))
        raise


def _extract_via_temp_copy(source: Path, *, preferred_name: str | None = None) -> str:
    """Copy ``source`` to temp, extract, delete the copy. Original file is never opened for write."""
    suffix = source.suffix.lower()
    label = (preferred_name or source.name or "document").replace("\\", "/").split("/")[-1]
    # Preserve extension for MarkItDown/Magika; isolate under a unique temp dir.
    with tempfile.TemporaryDirectory(prefix="duckclaw_doc_") as tmp_dir:
        dest = Path(tmp_dir) / (label if Path(label).suffix.lower() == suffix else f"doc{suffix}")
        shutil.copyfile(source, dest)
        return _convert_path(dest)


def convert_bytes_to_text(*, data: bytes, filename: str) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix not in EXTRACT_SUFFIXES:
        raise ValueError(f"MarkItDown solo admite: {', '.join(sorted(EXTRACT_SUFFIXES))}")
    safe_name = Path(filename).name or f"doc{suffix}"
    with tempfile.TemporaryDirectory(prefix="duckclaw_doc_") as tmp_dir:
        tmp_path = Path(tmp_dir) / safe_name
        if tmp_path.suffix.lower() != suffix:
            tmp_path = Path(tmp_dir) / f"doc{suffix}"
        tmp_path.write_bytes(data)
        return _convert_path(tmp_path)


def convert_file_path_to_text(path: str | Path) -> str:
    target = Path(path).expanduser().resolve()
    suffix = target.suffix.lower()
    if suffix in INGEST_NATIVE_SUFFIXES:
        # Plain text: read-only; no need to copy.
        return target.read_text(encoding="utf-8", errors="replace").strip()
    if suffix not in EXTRACT_SUFFIXES:
        if suffix in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"}:
            raise ValueError(
                f"extract_document_text no admite extensión: {suffix}. "
                "Capturas/imágenes: visión (VLM) o Gmail MCP; MarkItDown solo PDF/Office/HTML."
            )
        raise ValueError(f"extract_document_text no admite extensión: {suffix}")
    return _extract_via_temp_copy(target)


def extract_document_text_from_path(path: str | Path) -> tuple[str, str]:
    """Return (text, mime_hint). Operates on a temp copy for binary formats."""
    target = Path(path).expanduser().resolve()
    if not target.is_file():
        raise ValueError(f"No existe el archivo: {target.name}")
    text = convert_file_path_to_text(target)
    if not text.strip():
        raise ValueError(f"No se extrajo texto de: {target.name}")
    suffix = target.suffix.lower()
    if suffix in EXTRACT_SUFFIXES:
        return text, "text/plain"
    if suffix in {".md", ".markdown"}:
        return text, "text/markdown"
    return text, "text/plain"
